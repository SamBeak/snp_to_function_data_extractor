import requests
import json
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
from datetime import datetime
import time
import os

species = 'GCF_000003055.6'
api_key = os.getenv('API_KEY')

# 설정값
PROGRESS_FILE = 'progress_ncbi.json'
EXCEL_FILE = 'ncbi_gene_data_output.xlsx'
AUTO_SAVE_INTERVAL = 10  # 1000개 처리마다 자동 저장
MAX_CONSECUTIVE_FAILURES = 3  # 연속 실패 허용 횟수

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('gene_automation_ncbi.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def get_annotation_report(accession):
    url = f"https://api.ncbi.nlm.nih.gov/datasets/v2/genome/accession/{accession}/annotation_report"
    headers = {
        "X-Api-Key": api_key,
    }

    # Rate limiting: NCBI 요청 간 0.5초 대기 (NCBI 정책 준수)
    time.sleep(0.5)

    # 재시도 로직 (최대 3번)
    max_retries = 3
    for attempt in range(max_retries):
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                if data:
                    return data
                logger.warning(f'  └─ NCBI API 응답이 비어있음')
                return None
            elif resp.status_code == 429:  # Too Many Requests
                logger.warning(f'  └─ NCBI Rate limit 도달, 5초 대기 후 재시도 ({attempt + 1}/{max_retries})')
                time.sleep(5)
            else:
                logger.warning(f'  └─ NCBI 응답 코드: {resp.status_code}, 재시도 ({attempt + 1}/{max_retries})')
                time.sleep(2)
        except requests.exceptions.Timeout:
            logger.warning(f'  └─ NCBI 요청 타임아웃, 재시도 ({attempt + 1}/{max_retries})')
            time.sleep(2)
        except requests.exceptions.ConnectionError as e:
            logger.warning(f'  └─ NCBI 연결 에러 (네트워크 문제): {e}')
            time.sleep(2)
        except requests.exceptions.RequestException as e:
            logger.warning(f'  └─ NCBI 요청 에러: {e}, 재시도 ({attempt + 1}/{max_retries})')
            time.sleep(2)
        except Exception as e:
            logger.error(f'  └─ Annotation report 요청 중 예외 발생: {e}')
            time.sleep(2)

    # 모든 재시도 실패
    logger.error(f'  └─ Annotation report 요청 실패 (모든 재시도 소진)')
    return None


def get_function(gene_id):
    url = f"https://api.ncbi.nlm.nih.gov/datasets/v2/gene/id/{gene_id}"
    headers = {
        "X-Api-Key": api_key,
    }

    # Rate limiting: NCBI 요청 간 0.5초 대기 (NCBI 정책 준수)
    time.sleep(0.5)

    # 재시도 로직 (최대 3번)
    max_retries = 3
    for attempt in range(max_retries):
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                if data:
                    # reports 배열에서 gene_ontology 추출
                    reports = data.get("reports", [])
                    if not reports or not isinstance(reports, list):
                        return []

                    # 첫 번째 report의 gene 정보
                    gene = reports[0].get("gene")
                    if gene is None:
                        return []

                    # gene_ontology에서 molecular_functions 추출
                    gene_ontology = gene.get("gene_ontology")
                    if gene_ontology is None:
                        return []

                    molecular_functions = gene_ontology.get("molecular_functions", [])
                    if not isinstance(molecular_functions, list):
                        return []

                    # name 필드만 추출
                    function_names = []
                    for func in molecular_functions:
                        if isinstance(func, dict):
                            name = func.get("name")
                            if name:
                                function_names.append(name)

                    return function_names
                logger.warning(f'  └─ Gene Function API 응답이 비어있음')
                return []
            elif resp.status_code == 429:  # Too Many Requests
                logger.warning(f'  └─ NCBI Rate limit 도달, 5초 대기 후 재시도 ({attempt+1}/{max_retries})')
                time.sleep(5)
            else:
                logger.warning(f'  └─ NCBI 응답 코드: {resp.status_code}, 재시도 ({attempt+1}/{max_retries})')
                time.sleep(2)
        except requests.exceptions.Timeout:
            logger.warning(f'  └─ NCBI 요청 타임아웃, 재시도 ({attempt+1}/{max_retries})')
            time.sleep(2)
        except requests.exceptions.ConnectionError as e:
            logger.warning(f'  └─ NCBI 연결 에러 (네트워크 문제): {e}')
            time.sleep(2)
        except requests.exceptions.RequestException as e:
            logger.warning(f'  └─ NCBI 요청 에러: {e}, 재시도 ({attempt+1}/{max_retries})')
            time.sleep(2)
        except Exception as e:
            logger.error(f'  └─ Gene Function 요청 중 예외 발생: {e}')
            time.sleep(2)

    # 모든 재시도 실패
    logger.error(f'  └─ Gene Function 요청 실패 (모든 재시도 소진)')
    return []


def load_positions_from_json(file_path):
    """snps.json 파일에서 위치 정보를 읽어옴"""
    with open(file_path, 'r') as f:
        snps = json.load(f).get('snps')

    if snps is None:
        return []
    # "11: 55704515" 형식을 파싱
    parsed_snps = []
    for snp in snps:
        chrom_value, position_value = snp.split(':')
        position_value = position_value.strip()
        parsed_snps.append({
            'chrom': chrom_value,
            'pos': int(position_value)  # int로 변환
        })
    return parsed_snps


def save_progress(current_index, total_count, wb):
    """현재 진행 상황을 저장"""
    try:
        # Excel 파일 저장
        wb.save(EXCEL_FILE)
        logger.info(f'중간 저장 완료: {EXCEL_FILE}')

        # 진행 상황 저장
        progress_data = {
            'last_processed_index': current_index,
            'total_count': total_count,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        with open(PROGRESS_FILE, 'w') as f:
            json.dump(progress_data, f, indent=2)
        logger.info(f'진행 상황 저장: {current_index}/{total_count}')
        return True
    except Exception as e:
        logger.error(f'저장 실패: {e}')
        return False


def load_progress():
    """이전 진행 상황을 불러옴"""
    if not os.path.exists(PROGRESS_FILE):
        return None

    try:
        with open(PROGRESS_FILE, 'r') as f:
            progress = json.load(f)
        logger.info(
            f'이전 진행 상황 발견: {progress["last_processed_index"]}/{progress["total_count"]} ({progress["timestamp"]})')
        return progress
    except Exception as e:
        logger.error(f'진행 상황 로드 실패: {e}')
        return None


def handle_network_error(consecutive_failures):
    """네트워크 에러 처리"""
    if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
        logger.warning(f'\n{"=" * 60}')
        logger.warning(f'연속 {MAX_CONSECUTIVE_FAILURES}회 네트워크 에러 발생')
        logger.warning(f'네트워크 연결을 확인하고 Enter 키를 눌러 재개하세요...')
        logger.warning(f'{"=" * 60}\n')
        input()  # 사용자 입력 대기
        return 0  # 실패 카운터 리셋
    return consecutive_failures


snps_value = load_positions_from_json('snps.json')

# 이전 진행 상황 확인
previous_progress = load_progress()
start_index = 0
wb = None
ws = None

if previous_progress and os.path.exists(EXCEL_FILE):
    # 이전 작업 이어서 진행
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        start_index = previous_progress['last_processed_index'] + 1
        logger.info(f'이전 작업을 이어서 진행합니다. 시작 인덱스: {start_index}')

        # 재개 여부 확인
        response = input(f'\n이전 진행 상황에서 재개하시겠습니까? (y/n): ')
        if response.lower() != 'y':
            logger.info('처음부터 새로 시작합니다.')
            start_index = 0
            wb = Workbook()
            ws = wb.active
            ws.title = "Gene Data"
            # 헤더 작성
            ws['A1'] = 'SNP'
            ws['B1'] = 'GeneID'
            ws['C1'] = 'Gene'
            ws['D1'] = 'Function'
    except Exception as e:
        logger.error(f'이전 파일 로드 실패: {e}. 새로 시작합니다.')
        start_index = 0
        wb = Workbook()
        ws = wb.active
        ws.title = "Gene Data"
        # 헤더 작성
        ws['A1'] = 'SNP'
        ws['B1'] = 'GeneID'
        ws['C1'] = 'Gene'
        ws['D1'] = 'Function'
else:
    # 새로 시작
    wb = Workbook()
    ws = wb.active
    ws.title = "Gene Data"
    # 헤더 작성
    ws['A1'] = 'SNP'
    ws['B1'] = 'GeneID'
    ws['C1'] = 'Gene'
    ws['D1'] = 'Function'

# snps_value가 빈 값이 아니면
if snps_value:
    total_count = len(snps_value)
    logger.info(f'=== 유전자 데이터 처리 시작 ===')
    logger.info(f'총 처리할 SNP 개수: {total_count}')
    if start_index > 0:
        logger.info(f'시작 위치: {start_index} (남은 개수: {total_count - start_index})')

    row = 2 + start_index  # 데이터는 2행부터 시작 + 이전 진행 상황
    consecutive_failures = 0  # 연속 실패 카운터

    for i in range(start_index, len(snps_value)):
        chrom = snps_value[i]['chrom']
        pos = snps_value[i]['pos']
        snp_value = str(chrom) + ':' + str(pos)

        # 진행률 계산
        current_index = i + 1
        progress_percent = (current_index / total_count) * 100
        logger.info(f'[{current_index}/{total_count}] ({progress_percent:.1f}%) 처리 중: {snp_value}')

        result_genes = []

        # API 요청 실행
        report = get_annotation_report(species)

        # NoneType 에러 방지: report가 None인 경우 처리
        if report is None:
            logger.error(f'  └─ API 응답 실패 - 데이터를 가져올 수 없음')
            consecutive_failures += 1

            # 중간 저장
            if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                save_progress(i - 1, total_count, wb)
                consecutive_failures = handle_network_error(consecutive_failures)

            # 빈 값으로 채우기
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = ''
            ws[f'C{row}'] = ''
            ws[f'D{row}'] = ''
            row += 1

            # 자동 저장
            if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                save_progress(i, total_count, wb)
            continue

        # reports 필드 안전하게 접근
        reports = report.get("reports")
        if reports is None or not isinstance(reports, list):
            logger.warning(f'  └─ reports 필드가 없거나 리스트가 아님')
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = ''
            ws[f'C{row}'] = ''
            ws[f'D{row}'] = ''
            row += 1
            consecutive_failures = 0
            if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                save_progress(i, total_count, wb)
            continue

        # 염색체 위치에 해당하는 유전자 찾기
        for report_item in reports:
            # 안전한 딕셔너리 접근
            annotation = report_item.get("annotation")
            if annotation is None:
                continue

            # chromosomes 필드 확인 (리스트)
            chromosomes = annotation.get("chromosomes", [])
            if not isinstance(chromosomes, list):
                chromosomes = []

            # 현재 SNP의 chromosome이 이 유전자의 chromosomes 리스트에 있는지 확인
            if chrom not in chromosomes:
                continue

            # genomic_regions에서 위치 정보 추출
            genomic_regions = annotation.get("genomic_regions", [])
            if not isinstance(genomic_regions, list):
                continue

            # 위치가 범위 내에 있는지 확인
            position_match = False
            for region in genomic_regions:
                gene_range = region.get("gene_range")
                if gene_range is None:
                    continue

                ranges = gene_range.get("range", [])
                if not isinstance(ranges, list):
                    continue

                for rng in ranges:
                    if not isinstance(rng, dict):
                        continue

                    begin_str = rng.get("begin", "0")
                    end_str = rng.get("end", "0")

                    try:
                        begin = int(begin_str)
                        end = int(end_str)

                        if begin <= pos <= end:
                            position_match = True
                            break
                    except (ValueError, TypeError):
                        continue

                if position_match:
                    break

            if position_match:
                result_genes.append(annotation)

        if not result_genes:
            logger.warning(f'  └─ 해당 위치에 유전자 정보 없음')
            # 빈 값으로 채우기
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = ''
            ws[f'C{row}'] = ''
            ws[f'D{row}'] = ''
            row += 1
            consecutive_failures = 0  # 성공적으로 처리됨
            # 자동 저장
            if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                save_progress(i, total_count, wb)
            continue
        else:
            for gene in result_genes:
                # 안전한 딕셔너리 접근
                gene_id = gene.get("gene_id", "")
                gene_symbol = gene.get("symbol", "")

                logger.info(f'  └─ Gene ID: {gene_id}, Symbol: {gene_symbol}')

                # Gene Function 조회
                functions = []
                if gene_id:
                    functions = get_function(gene_id)

                # Function 리스트를 콤마로 연결
                if functions:
                    functionStr = ", ".join(functions)
                    logger.info(f'  └─ Function 정보 {len(functions)}개 수집 완료')
                else:
                    functionStr = ""
                    logger.warning(f'  └─ Function 정보 없음')

                # Excel에 데이터 작성
                ws[f'A{row}'] = snp_value
                ws[f'B{row}'] = gene_id
                ws[f'C{row}'] = gene_symbol
                ws[f'D{row}'] = functionStr
                row += 1
                consecutive_failures = 0  # 성공적으로 처리됨

            # 자동 저장
            if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                save_progress(i, total_count, wb)
    # 최종 저장
    logger.info(f'=== 모든 SNP 처리 완료 ===')
    save_progress(len(snps_value) - 1, total_count, wb)
else:
    logger.error('snp.json 파일 형식에 오류가 있습니다.')
    if wb:
        wb.close()
    exit()

# 최종 Excel 파일 저장
wb.save(EXCEL_FILE)
logger.info(f'Excel 파일 최종 저장 완료: {EXCEL_FILE}')

# 진행 상황 파일 삭제 (완료되었으므로)
if os.path.exists(PROGRESS_FILE):
    os.remove(PROGRESS_FILE)
    logger.info(f'진행 상황 파일 삭제: {PROGRESS_FILE}')

wb.close()