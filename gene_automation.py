import requests
import json
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
from datetime import datetime
import time
import os

species = 'bos_taurus'

# 설정값
PROGRESS_FILE = 'progress.json'
EXCEL_FILE = 'gene_data_output.xlsx'
AUTO_SAVE_INTERVAL = 10  # 10개 처리마다 자동 저장
MAX_CONSECUTIVE_FAILURES = 3  # 연속 실패 허용 횟수

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('gene_automation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def get_gene_at_pos(species, chrom, pos):
    url = f'https://rest.ensembl.org/overlap/region/{species}/{chrom}:{pos}-{pos}'
    headers = {'Content-Type': 'application/json'}
    params = {'feature': 'gene'}
    try:
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        if resp.status_code != 200 or not resp.json():
            return None
        return resp.json()[0]
    except Exception as e:
        logger.error(f'  └─ Ensembl API 요청 실패: {e}')
        return None

def get_go_terms(gene_id):
    url = f'https://rest.ensembl.org/xrefs/id/{gene_id}'
    headers = {'Content-Type': 'application/json'}
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        return [ref['primary_id'] for ref in resp.json() if ref['dbname'] == 'GO']
    except Exception as e:
        logger.error(f'  └─ GO terms 요청 실패: {e}')
        return []

def get_go_description(go_id):
    url = f"https://rest.ensembl.org/ontology/id/{go_id}"
    headers = {'Content-Type': 'application/json'}
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 200:
            js = resp.json()
            # 1. label(이름)이 있으면 우선 반환
            label = js.get('label', '')
            # 2. label이 없으면 description(내용) 반환
            if label:
                return label
            desc = js.get('description', '')
            if desc:
                return desc
            # 3. 둘 다 없으면 GO term ID 반환
            return go_id
        return go_id
    except Exception as e:
        logger.error(f'  └─ GO description 요청 실패: {e}')
        return go_id


def load_positions_from_json(file_path):
    """snps.json 파일에서 위치 정보를 읽어옴"""
    with open(file_path, 'r') as f:
        snps = json.load(f).get('snps')

    if snps is None:
        return []
    # "11: 55704515" 형식을 파싱
    parsed_snps = []
    for snp in snps:
        chrom_value, position_value = snp.split(':')
        position_value = position_value.strip()
        parsed_snps.append({
            'chrom': chrom_value,
            'pos': position_value
        })
    return parsed_snps

def save_progress(current_index, total_count, wb):
    """현재 진행 상황을 저장"""
    try:
        # Excel 파일 저장
        wb.save(EXCEL_FILE)
        logger.info(f'중간 저장 완료: {EXCEL_FILE}')

        # 진행 상황 저장
        progress_data = {
            'last_processed_index': current_index,
            'total_count': total_count,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        with open(PROGRESS_FILE, 'w') as f:
            json.dump(progress_data, f, indent=2)
        logger.info(f'진행 상황 저장: {current_index}/{total_count}')
        return True
    except Exception as e:
        logger.error(f'저장 실패: {e}')
        return False

def load_progress():
    """이전 진행 상황을 불러옴"""
    if not os.path.exists(PROGRESS_FILE):
        return None

    try:
        with open(PROGRESS_FILE, 'r') as f:
            progress = json.load(f)
        logger.info(f'이전 진행 상황 발견: {progress["last_processed_index"]}/{progress["total_count"]} ({progress["timestamp"]})')
        return progress
    except Exception as e:
        logger.error(f'진행 상황 로드 실패: {e}')
        return None

def handle_network_error(consecutive_failures):
    """네트워크 에러 처리"""
    if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
        logger.warning(f'\n{"="*60}')
        logger.warning(f'연속 {MAX_CONSECUTIVE_FAILURES}회 네트워크 에러 발생')
        logger.warning(f'네트워크 연결을 확인하고 Enter 키를 눌러 재개하세요...')
        logger.warning(f'{"="*60}\n')
        input()  # 사용자 입력 대기
        return 0  # 실패 카운터 리셋
    return consecutive_failures


snps_value = load_positions_from_json('snps.json')

# 이전 진행 상황 확인
previous_progress = load_progress()
start_index = 0
wb = None
ws = None

if previous_progress and os.path.exists(EXCEL_FILE):
    # 이전 작업 이어서 진행
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        start_index = previous_progress['last_processed_index'] + 1
        logger.info(f'이전 작업을 이어서 진행합니다. 시작 인덱스: {start_index}')

        # 재개 여부 확인
        response = input(f'\n이전 진행 상황에서 재개하시겠습니까? (y/n): ')
        if response.lower() != 'y':
            logger.info('처음부터 새로 시작합니다.')
            start_index = 0
            wb = Workbook()
            ws = wb.active
            ws.title = "Gene Data"
            # 헤더 작성
            ws['A1'] = 'SNP'
            ws['B1'] = 'GeneID'
            ws['C1'] = 'Gene'
            ws['D1'] = 'Function'
    except Exception as e:
        logger.error(f'이전 파일 로드 실패: {e}. 새로 시작합니다.')
        start_index = 0
        wb = Workbook()
        ws = wb.active
        ws.title = "Gene Data"
        # 헤더 작성
        ws['A1'] = 'SNP'
        ws['B1'] = 'GeneID'
        ws['C1'] = 'Gene'
        ws['D1'] = 'Function'
else:
    # 새로 시작
    wb = Workbook()
    ws = wb.active
    ws.title = "Gene Data"
    # 헤더 작성
    ws['A1'] = 'SNP'
    ws['B1'] = 'GeneID'
    ws['C1'] = 'Gene'
    ws['D1'] = 'Function'

# snps_value가 빈 값이 아니면
if snps_value:
    total_count = len(snps_value)
    logger.info(f'=== 유전자 데이터 처리 시작 ===')
    logger.info(f'총 처리할 SNP 개수: {total_count}')
    if start_index > 0:
        logger.info(f'시작 위치: {start_index} (남은 개수: {total_count - start_index})')

    row = 2 + start_index  # 데이터는 2행부터 시작 + 이전 진행 상황
    consecutive_failures = 0  # 연속 실패 카운터

    for i in range(start_index, len(snps_value)):
        chrom = snps_value[i]['chrom']
        pos = snps_value[i]['pos']
        snp_value = chrom + ':' + pos

        # 진행률 계산
        current_index = i + 1
        progress_percent = (current_index / total_count) * 100
        logger.info(f'[{current_index}/{total_count}] ({progress_percent:.1f}%) 처리 중: {snp_value}')

        # 1. 유전자 정보
        gene = get_gene_at_pos(species, chrom, pos)
        if not gene:
            logger.warning(f'  └─ 해당 위치에 유전자 정보 없음')
            # 빈 값으로 채우기
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = ''
            ws[f'C{row}'] = ''
            ws[f'D{row}'] = ''
            row += 1
            consecutive_failures = 0  # 성공적으로 처리됨
            # 자동 저장
            if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                save_progress(i, total_count, wb)
            continue
        # 2. GO term ID 얻기
        go_terms = get_go_terms(gene['id'])
        logger.info(f'  └─ Gene ID: {gene["id"]}, Gene Symbol: {gene.get("external_name", "-")}')

        # NCBI GeneID(Entrez)와 페이지 URL 예시 (소, CTNNA2: 527492)
        if not gene['description']:
            logger.warning(f'  └─ Gene description 없음')
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = gene['id']
            ws[f'C{row}'] = gene.get('external_name', '-')
            ws[f'D{row}'] = ''
            row += 1
            consecutive_failures = 0  # 성공적으로 처리됨
            # 자동 저장
            if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                save_progress(i, total_count, wb)
            continue

        # NCBI 요청 시 예외 처리 및 재시도 로직 추가
        functions = []
        try:
            gene_ncbi_id = gene['description'].split('Acc:')[1].replace(']', '')
            url = f'https://www.ncbi.nlm.nih.gov/gene/{gene_ncbi_id}'
            headers = {'User-Agent': 'Mozilla/5.0'}

            # Rate limiting: NCBI 요청 간 0.5초 대기 (NCBI 정책 준수)
            time.sleep(0.5)

            # 재시도 로직 (최대 3번)
            max_retries = 3
            request_success = False
            for attempt in range(max_retries):
                try:
                    r = requests.get(url, headers=headers, timeout=30)
                    if r.status_code == 200:
                        request_success = True
                        break
                    elif r.status_code == 429:  # Too Many Requests
                        logger.warning(f'  └─ NCBI Rate limit 도달, 5초 대기 후 재시도 ({attempt+1}/{max_retries})')
                        time.sleep(5)
                    else:
                        logger.warning(f'  └─ NCBI 응답 코드: {r.status_code}, 재시도 ({attempt+1}/{max_retries})')
                        time.sleep(2)
                except requests.exceptions.Timeout:
                    logger.warning(f'  └─ NCBI 요청 타임아웃, 재시도 ({attempt+1}/{max_retries})')
                    time.sleep(2)
                except requests.exceptions.ConnectionError as e:
                    logger.warning(f'  └─ NCBI 연결 에러 (네트워크 문제): {e}')
                    consecutive_failures += 1
                    # 중간 저장
                    save_progress(i - 1, total_count, wb)
                    # 네트워크 에러 처리
                    consecutive_failures = handle_network_error(consecutive_failures)
                    time.sleep(2)
                except requests.exceptions.RequestException as e:
                    logger.warning(f'  └─ NCBI 요청 에러: {e}, 재시도 ({attempt+1}/{max_retries})')
                    time.sleep(2)

            if not request_success:
                # 모든 재시도 실패
                logger.error(f'  └─ NCBI 요청 실패 (모든 재시도 소진)')
                ws[f'A{row}'] = snp_value
                ws[f'B{row}'] = gene['id']
                ws[f'C{row}'] = gene.get('external_name', '-')
                ws[f'D{row}'] = ''
                row += 1
                consecutive_failures += 1
                # 자동 저장
                if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                    save_progress(i, total_count, wb)
                continue

            soup = BeautifulSoup(r.text, 'html.parser')

            # Gene Ontology(Molecular function) 테이블 파싱
            # 테이블 내 각 'Function' (혹은 'Molecular function') 행의 텍스트를 추출
            for tr in soup.find_all('tr'):
                # Function 또는 Molecular function 필드에 해당하는 행 판별
                if tr.text.strip().lower().startswith(
                        'enables') or 'binding' in tr.text.lower() or 'structural molecule activity' in tr.text.lower():
                    # 각 셀(열) 분리 (Function label만 추출)
                    tds = tr.find_all('td')
                    if tds:
                        # 첫 번째 셀 혹은 전체 텍스트로 Function 설명 추출
                        func_text = tds[0].text.strip()
                        functions.append(func_text)
                    else:
                        # td가 없으면 전체 텍스트(행)를 넣음
                        functions.append(tr.text.strip())
        except Exception as e:
            logger.error(f'  └─ NCBI 파싱 중 예외 발생: {e}')
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = gene['id']
            ws[f'C{row}'] = gene.get('external_name', '-')
            ws[f'D{row}'] = ''
            row += 1
            consecutive_failures += 1
            # 자동 저장
            if (i + 1) % AUTO_SAVE_INTERVAL == 0:
                save_progress(i, total_count, wb)
            continue

        if functions:
            functionStr = ""
            for func_idx in range(len(functions)):
                if func_idx < len(functions) - 1:
                    f = functions[func_idx] + ", "
                else:
                    f = functions[func_idx]
                functionStr += f
            logger.info(f'  └─ Function 정보 {len(functions)}개 수집 완료')
            # Excel에 데이터 작성
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = gene['id']
            ws[f'C{row}'] = gene.get('external_name', '-')
            ws[f'D{row}'] = functionStr
            row += 1
            consecutive_failures = 0  # 성공적으로 처리됨
        else:
            logger.warning(f'  └─ Function 정보 없음')
            ws[f'A{row}'] = snp_value
            ws[f'B{row}'] = gene['id']
            ws[f'C{row}'] = gene.get('external_name', '-')
            ws[f'D{row}'] = ''
            row += 1
            consecutive_failures = 0  # 성공적으로 처리됨

        # 자동 저장
        if (i + 1) % AUTO_SAVE_INTERVAL == 0:
            save_progress(i, total_count, wb)

    # 최종 저장
    logger.info(f'=== 모든 SNP 처리 완료 ===')
    save_progress(len(snps_value) - 1, total_count, wb)
else:
    logger.error('snp.json 파일 형식에 오류가 있습니다.')
    if wb:
        wb.close()
    exit()

# 최종 Excel 파일 저장
wb.save(EXCEL_FILE)
logger.info(f'Excel 파일 최종 저장 완료: {EXCEL_FILE}')

# 진행 상황 파일 삭제 (완료되었으므로)
if os.path.exists(PROGRESS_FILE):
    os.remove(PROGRESS_FILE)
    logger.info(f'진행 상황 파일 삭제: {PROGRESS_FILE}')

wb.close()

